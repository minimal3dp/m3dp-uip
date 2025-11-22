#!/usr/bin/env python3
"""
Extract formulas and data from Klipper Calibrations.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook("research/Klipper Calibrations.xlsx", data_only=False)

# Dictionary to store all sheet data
all_sheets = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_data = {"name": sheet_name, "cells": {}}

    # Iterate through all cells
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                cell_info = {
                    "value": str(cell.value) if cell.value else None,
                    "is_formula": cell.data_type == "f",
                }

                # If it's a formula, include the formula
                if cell.data_type == "f":
                    cell_info["formula"] = cell.value

                sheet_data["cells"][cell_ref] = cell_info

    all_sheets[sheet_name] = sheet_data

# Print sheet names
print("=" * 80)
print("AVAILABLE SHEETS:")
print("=" * 80)
for name in wb.sheetnames:
    print(f"  • {name}")

print("\n" + "=" * 80)
print("DETAILED SHEET CONTENTS:")
print("=" * 80)

# Print each sheet's contents
for sheet_name, data in all_sheets.items():
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")

    # Sort cells by row then column
    sorted_cells = sorted(
        data["cells"].items(),
        key=lambda x: (
            int("".join(filter(str.isdigit, x[0])) or 0),
            "".join(filter(str.isalpha, x[0])),
        ),
    )

    for cell_ref, info in sorted_cells:
        if info["is_formula"]:
            print(f"{cell_ref}: FORMULA = {info['formula']}")
        else:
            print(f"{cell_ref}: {info['value']}")
